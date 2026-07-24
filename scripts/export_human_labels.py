#!/usr/bin/env python3
"""从当前知识库导出实体对齐与 Claim 证据人工标注工作簿。"""
from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
HIGH_FILL = PatternFill("solid", fgColor="FCE4D6")
DONE_FILL = PatternFill("solid", fgColor="E2F0D9")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _rows(conn: sqlite3.Connection, sql: str) -> list[dict]:
    return [dict(row) for row in conn.execute(sql)]


def entity_rows(conn: sqlite3.Connection) -> list[dict]:
    alignments = _rows(conn, """
        SELECT
            'alignment:' || ac.id AS item_id,
            'suspected_alignment' AS item_kind,
            ac.observed_name,
            e.canonical_name AS target_name,
            COALESCE(o.subject_type, '') AS observed_type,
            e.entity_type AS target_type,
            ac.score,
            ac.evidence_count,
            ac.independent_sources,
            COALESCE(s.name, '') AS source,
            COALESCE(o.excerpt, '') AS evidence,
            COALESCE(re.reason, '') AS model_reason,
            'uncertain' AS suggested_label
        FROM entity_alignment_candidates ac
        JOIN entities e ON e.id=ac.entity_id
        LEFT JOIN entity_alignment_evidence ae
          ON ae.candidate_id=ac.id
        LEFT JOIN observations o ON o.id=ae.observation_id
        LEFT JOIN source_snapshots ss ON ss.id=ae.source_snapshot_id
        LEFT JOIN sources s ON s.id=ss.source_id
        LEFT JOIN entity_resolution_events re ON re.observation_id=o.id
        WHERE ac.status='suspected'
        GROUP BY ac.id
        ORDER BY ac.score DESC,ac.id
    """)
    aliases = _rows(conn, """
        SELECT
            'alias:' || a.id AS item_id,
            'proposed_alias' AS item_kind,
            a.name AS observed_name,
            e.canonical_name AS target_name,
            '' AS observed_type,
            e.entity_type AS target_type,
            NULL AS score,
            1 AS evidence_count,
            CASE WHEN s.independence_group IS NULL THEN 0 ELSE 1 END
              AS independent_sources,
            COALESCE(s.name, '') AS source,
            a.evidence_excerpt AS evidence,
            '' AS model_reason,
            CASE
              WHEN REPLACE(a.normalized_name,'`','')='ce'
                THEN 'different_entity'
              WHEN a.normalized_name LIKE '%softmax%交叉熵%'
                THEN 'different_entity'
              ELSE 'same_entity'
            END AS suggested_label
        FROM aliases a
        JOIN entities e ON e.id=a.entity_id
        LEFT JOIN source_snapshots ss ON ss.id=a.source_snapshot_id
        LEFT JOIN sources s ON s.id=ss.source_id
        WHERE a.status='proposed'
          AND NOT EXISTS (
            SELECT 1 FROM entity_alignment_candidates ac
            WHERE ac.normalized_name=a.normalized_name
              AND ac.entity_id=a.entity_id
              AND ac.status='suspected'
          )
        ORDER BY a.id
    """)
    return alignments + aliases


def claim_rows(conn: sqlite3.Connection, limit: int) -> list[dict]:
    return _rows(conn, f"""
        WITH latest AS (
          SELECT target_id,MAX(id) decision_id
          FROM decisions
          WHERE target_type='claim'
          GROUP BY target_id
        )
        SELECT
          ev.id AS evidence_id,
          c.id AS claim_id,
          se.canonical_name AS subject,
          se.entity_type AS subject_type,
          c.relation,
          oe.canonical_name AS object,
          oe.entity_type AS object_type,
          COALESCE(src.name,'') AS source,
          ev.evidence_type,
          ev.location,
          ev.excerpt,
          ev.entailment AS model_label,
          COALESCE(json_extract(ev.metadata,'$.entailment_reason'),'')
            AS model_reason,
          COALESCE(d.outcome,'') AS claim_decision,
          COALESCE(d.reason,'') AS decision_reason,
          CASE
            WHEN ev.id IN (55,67) THEN 'high'
            WHEN ev.entailment IN ('contradicts','insufficient') THEN 'high'
            WHEN d.outcome='human_review' THEN 'high'
            ELSE 'medium'
          END AS priority,
          CASE
            WHEN ev.id=55 THEN 'supports'
            WHEN ev.id=67 THEN 'insufficient'
            ELSE ev.entailment
          END AS suggested_label
        FROM evidence ev
        JOIN claims c ON c.id=ev.claim_id
        JOIN entities se ON se.id=c.subject_id
        JOIN entities oe ON oe.id=c.object_id
        JOIN source_snapshots ss ON ss.id=ev.source_snapshot_id
        JOIN sources src ON src.id=ss.source_id
        LEFT JOIN latest l ON l.target_id=c.id
        LEFT JOIN decisions d ON d.id=l.decision_id
        ORDER BY
          CASE
            WHEN ev.id IN (55,67) THEN 0
            WHEN ev.entailment IN ('contradicts','insufficient') THEN 1
            WHEN d.outcome='human_review' THEN 2
            ELSE 3
          END,
          ev.id
        LIMIT {max(1, limit)}
    """)


def style_table(sheet, headers: list[str], input_columns: set[str]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 30
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if headers[cell.column - 1] in input_columns:
                cell.fill = INPUT_FILL
                cell.font = Font(color="0000FF")
        sheet.row_dimensions[row[0].row].height = 58
    label_col = headers.index("human_label") + 1
    letter = sheet.cell(1, label_col).column_letter
    sheet.conditional_formatting.add(
        f"{letter}2:{letter}{max(2, sheet.max_row)}",
        FormulaRule(formula=[f'LEN({letter}2)>0'], fill=DONE_FILL))


def add_entity_sheet(wb: Workbook, rows: list[dict]) -> None:
    sheet = wb.create_sheet("实体对齐")
    headers = [
        "item_id", "priority", "item_kind", "observed_name", "target_name",
        "observed_type", "target_type", "score", "evidence_count",
        "independent_sources", "source", "evidence", "model_reason",
        "suggested_label", "human_label", "human_canonical_name",
        "human_entity_type", "notes",
    ]
    sheet.append(headers)
    for item in rows:
        priority = (
            "high"
            if item["item_kind"] == "suspected_alignment"
            or item["suggested_label"] == "different_entity"
            else "medium")
        sheet.append([
            item["item_id"], priority, item["item_kind"], item["observed_name"],
            item["target_name"], item["observed_type"], item["target_type"],
            item["score"], item["evidence_count"], item["independent_sources"],
            item["source"], item["evidence"], item["model_reason"],
            item["suggested_label"], "", "", "", "",
        ])
    widths = {
        "A": 18, "B": 10, "C": 22, "D": 24, "E": 24, "F": 16, "G": 14,
        "H": 10, "I": 14, "J": 18, "K": 24, "L": 54, "M": 54, "N": 18,
        "O": 18, "P": 24, "Q": 18, "R": 32,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    style_table(
        sheet, headers,
        {"human_label", "human_canonical_name", "human_entity_type", "notes"})
    validation = DataValidation(
        type="list",
        formula1='"same_entity,different_entity,uncertain"',
        allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"O2:O{max(2, sheet.max_row)}")
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 2).value == "high":
            sheet.cell(row, 2).fill = HIGH_FILL
    sheet.sheet_properties.tabColor = "5B9BD5"


def add_claim_sheet(wb: Workbook, rows: list[dict]) -> None:
    sheet = wb.create_sheet("Claim证据")
    headers = [
        "evidence_id", "claim_id", "priority", "subject", "subject_type",
        "relation", "object", "object_type", "source", "evidence_type",
        "location", "excerpt", "model_label", "model_reason",
        "claim_decision", "decision_reason", "suggested_label",
        "human_label", "direct_support", "notes",
    ]
    sheet.append(headers)
    for item in rows:
        sheet.append([
            item["evidence_id"], item["claim_id"], item["priority"],
            item["subject"], item["subject_type"], item["relation"],
            item["object"], item["object_type"], item["source"],
            item["evidence_type"], item["location"], item["excerpt"],
            item["model_label"], item["model_reason"], item["claim_decision"],
            item["decision_reason"], item["suggested_label"], "", "", "",
        ])
    widths = {
        "A": 12, "B": 10, "C": 10, "D": 22, "E": 14, "F": 20, "G": 22,
        "H": 14, "I": 24, "J": 24, "K": 18, "L": 65, "M": 16, "N": 58,
        "O": 18, "P": 42, "Q": 18, "R": 18, "S": 16, "T": 32,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    style_table(sheet, headers, {"human_label", "direct_support", "notes"})
    label_validation = DataValidation(
        type="list",
        formula1='"supports,contradicts,insufficient,uncertain"',
        allow_blank=True)
    direct_validation = DataValidation(
        type="list", formula1='"yes,no,uncertain"', allow_blank=True)
    sheet.add_data_validation(label_validation)
    sheet.add_data_validation(direct_validation)
    label_validation.add(f"R2:R{max(2, sheet.max_row)}")
    direct_validation.add(f"S2:S{max(2, sheet.max_row)}")
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 3).value == "high":
            sheet.cell(row, 3).fill = HIGH_FILL
    sheet.sheet_properties.tabColor = "ED7D31"


def add_readme(wb: Workbook, entity_count: int, claim_count: int) -> None:
    sheet = wb.active
    sheet.title = "说明"
    sheet.sheet_properties.tabColor = "70AD47"
    content = [
        ("知识图谱人工标注种子集", "请只填写黄色列；不要修改 item_id/evidence_id。"),
        ("本次规模", f"实体对齐 {entity_count} 条；Claim 证据 {claim_count} 条。"),
        ("实体标签", "same_entity=同一实体或有效别名；different_entity=不同实体；uncertain=上下文不足。"),
        ("Claim 标签", "supports=原文直接支持；contradicts=原文明示相反；insufficient=相关但不能直接推出；uncertain=无法判断。"),
        ("直接性", "Claim 需要额外常识、多跳推理、上下位替换或方向转换时，direct_support 应选 no。"),
        ("标注原则", "按给出的来源原文判断，不使用模型建议替代人工判断；建议标签仅用于提高效率。"),
        ("回写", "标完后把文件交回 Codex；回写前会先校验 item_id、标签枚举和冲突。"),
    ]
    for key, value in content:
        sheet.append([key, value])
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 100
    for row in sheet.iter_rows():
        row[0].font = Font(bold=True, color="1F4E78")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row[0].row].height = 36
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = HEADER_FONT
    sheet["B1"].fill = HEADER_FILL
    sheet["B1"].font = HEADER_FONT
    sheet.freeze_panes = "A2"


def verify(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    if wb.sheetnames != ["说明", "实体对齐", "Claim证据"]:
        raise RuntimeError(f"工作表不完整: {wb.sheetnames}")
    errors = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    errors.append(f"{sheet.title}!{cell.coordinate}={cell.value}")
    if errors:
        raise RuntimeError("发现 Excel 错误: " + ", ".join(errors))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", default="data/kg.db")
    parser.add_argument(
        "--out", default="out/kg-human-labeling-20260724.xlsx")
    parser.add_argument("--claim-limit", type=int, default=30)
    args = parser.parse_args()

    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row
    entities = entity_rows(conn)
    claims = claim_rows(conn, args.claim_limit)
    conn.close()

    wb = Workbook()
    add_readme(wb, len(entities), len(claims))
    add_entity_sheet(wb, entities)
    add_claim_sheet(wb, claims)
    path = Path(args.out)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    verify(path)
    print(f"{path}: 实体对齐 {len(entities)} 条，Claim 证据 {len(claims)} 条")


if __name__ == "__main__":
    main()
